# Copyright (c) OpenMMLab. All rights reserved.
import argparse
import cv2
import mmcv
import numpy as np
import os
import os.path as osp
import shutil
import torch
import warnings
# import csv
import openpyxl
from pathlib import Path
from os import listdir
from os.path import isfile, join
from scipy.optimize import linear_sum_assignment
from PIL import ImageFont, ImageDraw, Image
import matplotlib.pyplot as plt
from sklearn.metrics import accuracy_score

from pyskl import __version__
from pyskl.apis import inference_recognizer, init_recognizer
from pyskl.smp import mrlines

try:
    from mmdet.apis import inference_detector, init_detector
except (ImportError, ModuleNotFoundError):
    def inference_detector(*args, **kwargs):
        pass

    def init_detector(*args, **kwargs):
        pass
    warnings.warn(
        'Failed to import `inference_detector` and `init_detector` from `mmdet.apis`. '
        'Make sure you can successfully import these if you want to use related features. '
    )

try:
    from mmpose.apis import inference_top_down_pose_model, init_pose_model, vis_pose_result
except (ImportError, ModuleNotFoundError):
    def init_pose_model(*args, **kwargs):
        pass

    def inference_top_down_pose_model(*args, **kwargs):
        pass

    def vis_pose_result(*args, **kwargs):
        pass

    warnings.warn(
        'Failed to import `init_pose_model`, `inference_top_down_pose_model`, `vis_pose_result` from '
        '`mmpose.apis`. Make sure you can successfully import these if you want to use related features. '
    )


try:
    import moviepy.editor as mpy
except ImportError:
    raise ImportError('Please install moviepy to enable output file')

FONTFACE = cv2.FONT_ITALIC #cv2.FONT_HERSHEY_DUPLEX
FONTSCALE = 0.75
FONTCOLOR = (0,0,255) # Red
THICKNESS = 1
LINETYPE = 1


def parse_args():
    parser = argparse.ArgumentParser(description='PoseC3D demo')
    # parser.add_argument('video', help='video file/url')
    # parser.add_argument('out_filename', help='output filename')
    parser.add_argument('--save-folder-path', default=None)
    parser.add_argument('--video-folder-path', default=None)
    # parser.add_argument('--img_folder_path', default=None)
    parser.add_argument(
        '--config',
        default='configs/posec3d/slowonly_r50_ntu120_xsub/joint.py',
        help='skeleton action recognition config file path')
    parser.add_argument(
        '--checkpoint',
        default='https://download.openmmlab.com/mmaction/pyskl/ckpt/posec3d/slowonly_r50_ntu120_xsub/joint.pth',
        help='skeleton action recognition checkpoint file/url')
    parser.add_argument(
        '--det-config',
        default='demo/faster_rcnn_r50_fpn_1x_coco-person.py',
        help='human detection config file path (from mmdet)')
    parser.add_argument(
        '--det-checkpoint',
        default=('https://download.openmmlab.com/mmdetection/v2.0/faster_rcnn/faster_rcnn_r50_fpn_1x_coco-person/'
                 'faster_rcnn_r50_fpn_1x_coco-person_20201216_175929-d022e227.pth'),
        help='human detection checkpoint file/url')
    parser.add_argument(
        '--pose-config',
        default='demo/hrnet_w32_coco_256x192.py',
        help='human pose estimation config file path (from mmpose)')
    parser.add_argument(
        '--pose-checkpoint',
        default='https://download.openmmlab.com/mmpose/top_down/hrnet/hrnet_w32_coco_256x192-c78dce93_20200708.pth',
        help='human pose estimation checkpoint file/url')
    parser.add_argument(
        '--det-score-thr',
        type=float,
        default=0.9,
        help='the threshold of human detection score')
    parser.add_argument(
        '--label-map',
        default='tools/data/label_map/nturgbd_120.txt',
        help='label map file')
    parser.add_argument(
        '--device', type=str, default='cuda:0', help='CPU/CUDA device option')
    parser.add_argument(
        '--short-side',
        type=int,
        default=480,
        help='specify the short-side length of the image')
    parser.add_argument('--video-action-annotations', type=str, help='text file path containing video path and corresponding action label')
    args = parser.parse_args()
    return args


class ExcelBuilder:
    def __init__(self, filename):
        self.workbook = openpyxl.Workbook()
        self.filename = filename

    def add_sheet(self, sheet_name):
        sheet = self.workbook.create_sheet(sheet_name)
        return sheet

    def write_data(self, sheet, data, start_row=1, start_col=1):
        for row_idx, row in enumerate(data):
            for col_idx, value in enumerate(row):
                sheet.cell(row=start_row + row_idx, column=start_col + col_idx).value = value

    def delete_sheet(self, sheet_name):
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
        else:
            raise KeyError(f"Sheet '{sheet_name}' does not exist in the workbook.")

    def save(self):
        self.workbook.save(self.filename)


def frame_extraction(video_path, short_side):
    """Extract frames given video_path.

    Args:
        video_path (str): The video_path.
    """
    # Load the video, extract frames into ./tmp/video_name
    target_dir = osp.join('./tmp', osp.basename(osp.splitext(video_path)[0]))
    os.makedirs(target_dir, exist_ok=True)
    
    # Should be able to handle videos up to several hours
    frame_tmpl = osp.join(target_dir, 'img_{:06d}.jpg')
    vid = cv2.VideoCapture(video_path)
    frames = []
    frame_paths = []
    flag, frame = vid.read()
    cnt = 0
    new_h, new_w = None, None

    while flag:
        if new_h is None:
            h, w, _ = frame.shape
            new_w, new_h = mmcv.rescale_size((w, h), (short_side, np.Inf))

        frame = mmcv.imresize(frame, (new_w, new_h))

        frames.append(frame)
        frame_path = frame_tmpl.format(cnt + 1)
        if cnt == 0:
            print('Sample frame save path:', frame_path)
        frame_paths.append(frame_path)

        cv2.imwrite(frame_path, frame)

        cnt += 1
        flag, frame = vid.read()

    return frame_paths, frames


def frame_extraction_images(img_folder_path, short_side):

    file_name_list = os.listdir(img_folder_path)
    file_name_list.sort()
    print(' ==> File name list:', file_name_list)
    
    frames = []
    frame_paths = []
    end = (".png", ".jpg")
    new_h, new_w = None, None
    for file_name in file_name_list:
    
        if (file_name.endswith(end)):
            frame_path = os.path.join(img_folder_path, file_name)
            frame_paths.append(frame_path)

            frame = cv2.imread(frame_path)
            if new_h is None:
                h, w, _ = frame.shape
                new_w, new_h = mmcv.rescale_size((w, h), (short_side, np.Inf)) 

            # img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            frame = mmcv.imresize(frame, (new_w, new_h))
            
            # plt.imshow(img, interpolation='nearest')
            # plt.show()
            frames.append(frame)
            
    return frame_paths, frames


def detection_inference(args, frame_paths):
    """Detect human boxes given frame paths.

    Args:
        args (argparse.Namespace): The arguments.
        frame_paths (list[str]): The paths of frames to do detection inference.

    Returns:
        list[np.ndarray]: The human detection results.
    """
    model = init_detector(args.det_config, args.det_checkpoint, args.device)
    assert model is not None, ('Failed to build the detection model. Check if you have installed mmcv-full properly. '
                               'You should first install mmcv-full successfully, then install mmdet, mmpose. ')
    assert model.CLASSES[0] == 'person', 'We require you to use a detector trained on COCO'
    results = []
    print('Performing Human Detection for each frame')

    prog_bar = mmcv.ProgressBar(len(frame_paths))
    for frame_path in frame_paths:
        result = inference_detector(model, frame_path)
        # We only keep human detections with score larger than det_score_thr
        result = result[0][result[0][:, 4] >= args.det_score_thr]
        results.append(result)
        prog_bar.update()
    
    print()

    return results


def pose_inference(args, frame_paths, det_results):
    model = init_pose_model(args.pose_config, args.pose_checkpoint,
                            args.device)
    ret = []
    print('Performing Human Pose Estimation for each frame')

    prog_bar = mmcv.ProgressBar(len(frame_paths))
    for f, d in zip(frame_paths, det_results):
        # Align input format
        d = [dict(bbox=x) for x in list(d)]
        pose = inference_top_down_pose_model(model, f, d, format='xyxy')[0]
        ret.append(pose)
        prog_bar.update()
    
    print()

    return ret


def dist_ske(ske1, ske2):
    dist = np.linalg.norm(ske1[:, :2] - ske2[:, :2], axis=1) * 2
    diff = np.abs(ske1[:, 2] - ske2[:, 2])
    return np.sum(np.maximum(dist, diff))


def pose_tracking(pose_results, max_tracks=2, thre=30):
    tracks, num_tracks = [], 0
    num_joints = None
    for idx, poses in enumerate(pose_results):
        if len(poses) == 0:
            continue
        if num_joints is None:
            num_joints = poses[0].shape[0]
        track_proposals = [t for t in tracks if t['data'][-1][0] > idx - thre]
        n, m = len(track_proposals), len(poses)
        scores = np.zeros((n, m))

        for i in range(n):
            for j in range(m):
                scores[i][j] = dist_ske(track_proposals[i]['data'][-1][1], poses[j])

        row, col = linear_sum_assignment(scores)
        for r, c in zip(row, col):
            track_proposals[r]['data'].append((idx, poses[c]))
        if m > n:
            for j in range(m):
                if j not in col:
                    num_tracks += 1
                    new_track = dict(data=[])
                    new_track['track_id'] = num_tracks
                    new_track['data'] = [(idx, poses[j])]
                    tracks.append(new_track)
    tracks.sort(key=lambda x: -len(x['data']))
    result = np.zeros((max_tracks, len(pose_results), num_joints, 3), dtype=np.float16)
    for i, track in enumerate(tracks[:max_tracks]):
        for item in track['data']:
            idx, pose = item
            result[i, idx] = pose
    return result[..., :2], result[..., 2]


def main():
    args = parse_args()
    print()

    ### Configs
    config = mmcv.Config.fromfile(args.config)
    config.data.test.pipeline = [x for x in config.data.test.pipeline if x['type'] != 'DecompressPose']

    if not os.path.exists(args.save_folder_path): 
        os.makedirs(args.save_folder_path)

    ### Init model
    model = init_recognizer(config, args.checkpoint, args.device)

    ### Load label_map
    label_map = [x.strip() for x in open(args.label_map).readlines()]  
    print(f' =====> Label map: {label_map}')
    
    ### Get paths of videos
    file_path_list = [f for f in listdir(args.video_folder_path) if isfile(join(args.video_folder_path, f))]
    print(f' =====> Num of videos: {len(file_path_list)}') 
    
        
    ### Label list
    label_dict = {}
    lines = mrlines(args.video_action_annotations)
    lines = [x.split() for x in lines]
    for i in range(len(lines)):
        video_name = lines[i][0]
        label = int(lines[i][1])
        # label_str = label_map[label]
        label_dict[Path(video_name).stem + '.mp4'] = label
    print(f' =====> Label Dict: {label_dict}')

    print()
    print()

    keys = [i for i in range(len(label_map))]
    predicted_labels_dict = {key: [] for key in keys}

    for i, video_name in enumerate(file_path_list):
        
        print(f' =====> Video {i+1}, name: {video_name}')
        label = label_dict[video_name]
        label_str = label_map[label]
        print(f' =====> Label Action: {label_str}, Idx: {label}')

        # if not args.img_folder_path is None:
        #     frame_paths, original_frames = frame_extraction_images(args.img_folder_path,
        #                                                     args.short_side)
        # else:
        #     frame_paths, original_frames = frame_extraction(args.video,
        #                                                     args.short_side)

        ### Extract frames
        frame_paths, original_frames = frame_extraction(os.path.join(args.video_folder_path, video_name),
                                                            args.short_side)

        num_frame = len(frame_paths)
        h, w, _ = original_frames[0].shape
        print(' =====> Build inference annotations ...')

        ### Human detection
        det_results = detection_inference(args, frame_paths)
        torch.cuda.empty_cache()

        ### Pose estimation
        pose_results = pose_inference(args, frame_paths, det_results)
        torch.cuda.empty_cache()

        ### Annotations dict
        fake_anno = dict(
            frame_dir='',
            label=-1,
            img_shape=(h, w),
            original_shape=(h, w),
            start_index=0,
            modality='Pose',
            total_frames=num_frame)

        num_person = max([len(x) for x in pose_results])
        num_keypoint = 17
        keypoint = np.zeros((num_person, num_frame, num_keypoint, 2),
                            dtype=np.float16)
        keypoint_score = np.zeros((num_person, num_frame, num_keypoint),
                                dtype=np.float16)
        for i, poses in enumerate(pose_results):
            for j, pose in enumerate(poses):
                pose = pose['keypoints']
                keypoint[j, i] = pose[:, :2]
                keypoint_score[j, i] = pose[:, 2]
        fake_anno['keypoint'] = keypoint
        fake_anno['keypoint_score'] = keypoint_score

        print(' =====> Build inference annotations ...')
        
        ### Action recognition
        results = inference_recognizer(model, fake_anno)
        print(f' =====> Top5 Predicts & Scores: {results}')

        predict = results[0][0]
        action_str = label_map[predict]
        print(f' =====> Top1 Predicted Action: {action_str}, Idx: {predict}')

        predicted_labels_dict[label].append(predict)
        print(f' =====> Pred Save Dict: {predicted_labels_dict}')

        ### Visualization
        pose_model = init_pose_model(args.pose_config, args.pose_checkpoint,
                                    args.device)
        vis_frames = [
            vis_pose_result(pose_model, frame_paths[i], pose_results[i])
            for i in range(num_frame)
        ]
        height, width, _ = np.array(vis_frames[0]).shape

        for i, frame in enumerate(vis_frames):
            # cv2.putText(frame, action_str, (5, 5), FONTFACE, FONTSCALE,
            #             FONTCOLOR, THICKNESS, LINETYPE)

            t_size = cv2.getTextSize(action_str, cv2.FONT_HERSHEY_PLAIN, 1, 1)[0]
            
            # Draw korean
            frame = Image.fromarray(frame)
            draw = ImageDraw.Draw(frame)

            red_color = (0, 0, 255) 
            font_size = 25
            
            # Korean font
            fontpath = "/usr/share/fonts/NanumFont/NanumGothic.ttf"
            font = ImageFont.truetype(fontpath, font_size)

            draw.text((5, 5), action_str, font=font, fill=red_color)
            
            frame = np.array(frame) 
            vis_frames[i] = frame

        ### Write video
        fourcc = cv2.VideoWriter_fourcc('m', 'p', '4', 'v')
        # writer = cv2.VideoWriter(args.out_filename, fourcc, fps=24, frameSize=(width, height))

        out_filename = os.path.join(args.save_folder_path, Path(video_name).stem + '_inf.mp4')
        writer = cv2.VideoWriter(out_filename, fourcc, fps=24, frameSize=(width, height))

        print(f' =====> Write video ...')
        # set_flag = False
        # for i, frame in enumerate(vis_frames): #enumerate([x[:, :, ::-1] for x in vis_frames]):
            
        #     # cv2.imwrite(f'./tmp/{i}.png', frame)
        #     # writer.write(frame)
			
        #     ax = plt.subplot(1,1,1)

        #     plt.ion()
            
        #     resize_width = 350
        #     resize_height = 350
        #     img = cv2.resize(frame, dsize=(resize_width, resize_height), interpolation=cv2.INTER_AREA)
        #     img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        #     # Open board
        #     if not set_flag:
        #         im = ax.imshow(img)
        #         set_flag = True

        #     # Change data
        #     im.set_data(img)

        #     # Wait
        #     plt.pause(0.0001)

        # plt.ioff()
        
        for i, frame in enumerate(vis_frames): #enumerate([x[:, :, ::-1] for x in vis_frames]):
            writer.write(frame)
        writer.release()

        print(f' =====> Complete to write video: {out_filename}')
        print()
        print()

        # vid = mpy.ImageSequenceClip(vis_frames, fps=24)
        # vid.write_videofile(args.out_filename, remove_temp=True, codec="libx264")

        # tmp_frame_dir = osp.dirname(frame_paths[0])
        # shutil.rmtree(tmp_frame_dir)

    print(' === Final accuracy each class ===')
    excel_file_path = os.path.join(args.save_folder_path, f'results.xlsx')
    builder = ExcelBuilder(excel_file_path)
    builder.delete_sheet('Sheet')

    for key in predicted_labels_dict.keys():

        sheet = builder.add_sheet(label_map[key])

        predicts = predicted_labels_dict[key]
        labels = [key for i in range(len(predicted_labels_dict[key]))]
        corrects = list(map(lambda x, y: 'O' if x == y else 'X', predicts, labels))
        
        accuracy = accuracy_score(labels, predicts)
        print(f' -- {label_map[key]} Class Accuracy: {accuracy * 100:.2f}%')
        
        sheet_data = [
            ['Labels'] + labels,
            ['Predicts'] + predicts,
            ['Corrects'] + corrects,
            ['Accuracy'] + [f'{accuracy * 100:.2f}%']
        ]   

        builder.write_data(sheet, sheet_data)

        # csv_file_path = os.path.join(args.save_folder_path, f'{key}_scores.csv')
        # with open(csv_file_path, 'w', newline='') as csvfile:
        #     csv_writer = csv.writer(csvfile)
        #     csv_writer.writerows(save_list)

    builder.save()
    print("Excel file created successfully!")
    
if __name__ == '__main__':
    main()
